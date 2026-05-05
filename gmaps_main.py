import os
import time
import re
import unicodedata
from datetime import date


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from gmaps_scraper import scrape_maps
from gmaps_web import scrape_website, deep_scrape_website

HISTORIA_PATH = 'historia_pobranych.txt'
RAPORTY_DIR = 'Raporty'


def wczytaj_historie(path: str) -> set:
    if not os.path.exists(path):
        return set()
    with open(path, encoding='utf-8') as f:
        return {line.strip() for line in f if line.strip()}


def dopisz_do_historii(path: str, nazwy: list):
    with open(path, 'a', encoding='utf-8') as f:
        for nazwa in nazwy:
            f.write(nazwa + '\n')


def wyczysc_historie(path: str):
    """Kasuje całą zawartość pliku historii."""
    with open(path, 'w', encoding='utf-8') as f:
        f.write('')


def usun_z_historii(path: str, nazwy_do_usuniecia: list):
    """Usuwa wybrane firmy z historii, pozostałe zostawia."""
    historia = wczytaj_historie(path)
    do_usuniecia = set(nazwy_do_usuniecia)
    pozostale = [n for n in historia if n not in do_usuniecia]
    with open(path, 'w', encoding='utf-8') as f:
        for nazwa in pozostale:
            f.write(nazwa + '\n')


def slugify(text: str) -> str:
    # Normalize unicode → ASCII, remove non-alphanumeric
    normalized = unicodedata.normalize('NFKD', text)
    ascii_text = normalized.encode('ascii', 'ignore').decode('ascii')
    slug = re.sub(r'[^\w\s]', '', ascii_text)
    slug = re.sub(r'\s+', '_', slug.strip())
    return slug[:50]  # cap length


def buduj_nazwe_pliku(query: str) -> str:
    slug = slugify(query)
    today = date.today().isoformat()
    base = os.path.join(RAPORTY_DIR, f'Wyniki_{slug}_{today}')
    path = base + '.xlsx'
    if not os.path.exists(path):
        return path
    v = 2
    while os.path.exists(f'{base}_v{v}.xlsx'):
        v += 1
    return f'{base}_v{v}.xlsx'


def zapisz_excel(firmy: list, path: str):
    columns = ['Nazwa Firmy', 'Adres', 'Telefon', 'Strona WWW', 'E-mail', 'NIP']
    rows = [
        [
            f['nazwa'],
            f['adres'],
            f['telefon'],
            f['www'],
            f['email'],
            f['nip'],
        ]
        for f in firmy
    ]
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(path, index=False)

    # Auto-fit column widths
    wb = load_workbook(path)
    ws = wb.active
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
    wb.save(path)


def zapisz_excel_linkedin(kontakty: list, path: str):
    columns = ['Imię i Nazwisko', 'Stanowisko', 'Firma', 'Lokalizacja', 'LinkedIn URL']
    rows = [
        [
            k.get('imie_nazwisko', ''),
            k.get('stanowisko', ''),
            k.get('firma', ''),
            k.get('lokalizacja', ''),
            k.get('linkedin_url', ''),
        ]
        for k in kontakty
    ]
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 80)
    wb.save(path)


def main():
    query = input("Podaj frazę wyszukiwania (np. 'Kancelaria prawna Poznań'): ").strip()
    if not query:
        print("Brak frazy. Zakończono.")
        return

    limit_input = input("Limit firm [domyślnie 20]: ").strip()
    limit = int(limit_input) if limit_input.isdigit() else 20

    historia = wczytaj_historie(HISTORIA_PATH)
    print(f"Historia: {len(historia)} znanych firm.")
    os.makedirs(RAPORTY_DIR, exist_ok=True)

    # Faza 1: Google Maps
    print(f"\nSzukam '{query}' na Google Maps (limit={limit})...")
    firmy = scrape_maps(query, limit=limit, already_collected=historia)
    print(f"\nZnaleziono {len(firmy)} nowych firm z Map.")

    if not firmy:
        print("Brak wyników. Zakończono.")
        return

    # Zapisz do historii od razu — zabezpiecza przed powtórnym pobraniem przy crashu w Fazie 2
    dopisz_do_historii(HISTORIA_PATH, [f['nazwa'] for f in firmy])
    print(f"Dodano {len(firmy)} firm do historii.")

    # Faza 2: Strony WWW
    print("\nScrapuję strony WWW firm...")
    for i, firma in enumerate(firmy):
        if firma['www']:
            dane = scrape_website(firma['www'])
            if not firma['email']:
                firma['email'] = dane['email']
            firma['nip'] = dane['nip']
        email_info = firma['email'] or 'brak emaila'
        print(f"  [{i+1}/{len(firmy)}] {firma['nazwa']} — {email_info}")
        time.sleep(1.5)

    # Faza 2b: Drugi przebieg dla firm z niekompletnymi danymi
    incomplete = [f for f in firmy if not f['email'] or not f['nip']]
    if incomplete:
        print(f"\n[Faza 2b] Głębsze szukanie dla {len(incomplete)} firm z brakującymi danymi...")
        for i, firma in enumerate(incomplete):
            dane = deep_scrape_website(firma['www'], firma['nazwa'])
            if dane['email'] and not firma['email']:
                firma['email'] = dane['email']
            if dane['nip'] and not firma['nip']:
                firma['nip'] = dane['nip']
            print(f"  [{i+1}/{len(incomplete)}] {firma['nazwa']} — "
                  f"email={firma['email'] or 'brak'} nip={firma['nip'] or 'brak'}")
            time.sleep(1.5)

    # Faza 3: Zapis Excel
    nazwa_pliku = buduj_nazwe_pliku(query)
    zapisz_excel(firmy, nazwa_pliku)
    print(f"\nZapisano: {nazwa_pliku}")



if __name__ == '__main__':
    main()
